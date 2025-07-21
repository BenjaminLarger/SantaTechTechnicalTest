import openpyxl
import pandas as pd
import matplotlib.pyplot as plt
import os
import datetime


# ============
wb_path = r"example_sheets/opos.xlsx"

# ============
workbook = openpyxl.load_workbook(wb_path)

# ============
# Load the 'Sheet1'
sheet = workbook["Sheet1"]

# Create a new sheet "Analysis"
analysis_sheet = workbook.create_sheet("Analysis")

# ============
import re

# Define today's date
today_date = datetime.datetime(2025, 6, 10)

# List to store cumulative row numbers
cumulative_rows = []

# Keywords for cumulative rows
cumulative_keywords = ['debitor', 'debtor', 'creditor']

# Loop through the rows to find cumulative rows
for row in range(2, 152):  # Start at row 2 to skip the header
    invoice_number = sheet.cell(row, 4).value
    due_date = sheet.cell(row, 16).value
    invoice_date = sheet.cell(row, 5).value
    amount = sheet.cell(row, 18).value
    is_cumulative = False

    # Check for empty values or a change in format detected by regex
    if (invoice_number is None or due_date is None or invoice_date is None or amount is None or
        not isinstance(invoice_date, datetime.datetime) or not isinstance(due_date, datetime.datetime) or
        any(keyword in str(invoice_number).lower() for keyword in cumulative_keywords)):
        is_cumulative = True

    # If the row is cumulative, add the row number to the list
    if is_cumulative:
        cumulative_rows.append(row)

# Print the list of cumulative rows for verification
print(cumulative_rows)

# Output cumulative rows to the 'Analysis' sheet
analysis_sheet.append(['Cumulative Rows'])
for row_num in cumulative_rows:
    analysis_sheet.append([row_num])

# ============
# List to store invoice row numbers
invoice_rows = []

# Loop through the rows to find invoice rows
for row in range(2, 152):
    if row not in cumulative_rows:  # Exclude cumulative rows
        amount = sheet.cell(row, 18).value
        if amount is not None and amount > 0:  # Check for positive invoice amount
            invoice_rows.append(row)

# Print the list of invoice rows for verification
print(invoice_rows)

# Output invoice rows to the 'Analysis' sheet
analysis_sheet.append(['Invoice Rows'])
for row_num in invoice_rows:
    analysis_sheet.append([row_num])

# ============
# List to store credit row numbers
credit_rows = []

# Loop through the rows to find credit rows
for row in range(2, 152):
    if row not in cumulative_rows:  # Exclude cumulative rows
        amount = sheet.cell(row, 18).value
        if amount is not None and amount < 0:  # Check for negative invoice amount
            credit_rows.append(row)

# Print the list of credit rows for verification
print(credit_rows)

# Output credit rows to the 'Analysis' sheet
analysis_sheet.append(['Credit Rows'])
for row_num in credit_rows:
    analysis_sheet.append([row_num])

# ============
# List to hold rows where information might be missing
missing_info_invoice = []
missing_info_credit = []

# Define the column numbers for the required information
required_columns = [4, 5, 16, 18]  # Columns: Belegnummer (D), Belegdatum (E), Nettofälligkeit (P), Betrag in Belegwährung (R)

# Function to check if required information is present
def is_information_present(row, required_columns):
    for column in required_columns:
        if sheet.cell(row, column).value is None:
            return False
    return True

# Check each "invoice" row for completeness
for row_num in invoice_rows:
    if not is_information_present(row_num, required_columns):
        missing_info_invoice.append(row_num)

# Check each "credit" row for completeness
for row_num in credit_rows:
    if not is_information_present(row_num, required_columns):
        missing_info_credit.append(row_num)

# Print the rows with missing information for invoices and credits
print(f'Invoice Rows Missing Information: {missing_info_invoice}')
print(f'Credit Rows Missing Information: {missing_info_credit}')

# Output missing info rows to the 'Analysis' sheet
analysis_sheet.append(['Missing Info Invoice Rows'])
for row_num in missing_info_invoice:
    analysis_sheet.append([row_num])

analysis_sheet.append(['Missing Info Credit Rows'])
for row_num in missing_info_credit:
    analysis_sheet.append([row_num])

# ============
# Sum the invoice amounts
invoice_amount_sum = sum(sheet.cell(row, 18).value for row in invoice_rows)

# Print the sum of the invoice amounts
print(f'Sum of Invoice Amounts: {invoice_amount_sum}')

# Output the sum of invoice amounts to the 'Analysis' sheet
analysis_sheet.append(['Sum of Invoice Amounts'])
analysis_sheet.append([invoice_amount_sum])

# ============
# Sum the credit amounts
credit_amount_sum = sum(sheet.cell(row, 18).value for row in credit_rows)

# Print the sum of the credit amounts
print(f'Sum of Credit Amounts: {credit_amount_sum}')

# Output the sum of credit amounts to the 'Analysis' sheet
analysis_sheet.append(['Sum of Credit Amounts'])
analysis_sheet.append([credit_amount_sum])

# ============
# Create clusters for ageing report
not_mature_invoices = 0
maturity_1_30_invoices = 0
maturity_31_60_invoices = 0
mature_over_60_invoices = 0

# Loop over invoice rows and calculate maturity to cluster invoice amounts
for row_num in invoice_rows:
    nettofalligkeit = sheet.cell(row_num, 16).value
    days_to_maturity = (nettofalligkeit - today_date).days
    invoice_amount = sheet.cell(row_num, 18).value
    if days_to_maturity > 0:
        not_mature_invoices += invoice_amount
    elif 1 <= days_to_maturity <= 30:
        maturity_1_30_invoices += invoice_amount
    elif 31 <= days_to_maturity <= 60:
        maturity_31_60_invoices += invoice_amount
    else:
        mature_over_60_invoices += invoice_amount

# Calculate the percentage for each cluster
total_invoice_amount = float(invoice_amount_sum)
percent_not_mature = (not_mature_invoices / total_invoice_amount) * 100
percent_maturity_1_30 = (maturity_1_30_invoices / total_invoice_amount) * 100
percent_maturity_31_60 = (maturity_31_60_invoices / total_invoice_amount) * 100
percent_mature_over_60 = (mature_over_60_invoices / total_invoice_amount) * 100

# Output the ageing report for invoice rows to the 'Analysis' sheet
analysis_sheet.append(['Ageing Report - Invoices'])
analysis_sheet.append(['Cluster', 'Amount', 'Percentage'])
analysis_sheet.append(['Not Mature', not_mature_invoices, percent_not_mature])
analysis_sheet.append(['1-30 Days', maturity_1_30_invoices, percent_maturity_1_30])
analysis_sheet.append(['31-60 Days', maturity_31_60_invoices, percent_maturity_31_60])
analysis_sheet.append(['>60 Days', mature_over_60_invoices, percent_mature_over_60])

# ============
# Create clusters for ageing report
not_mature_credits = 0
maturity_1_30_credits = 0
maturity_31_60_credits = 0
mature_over_60_credits = 0

# Loop over credit rows and calculate maturity to cluster credit amounts
for row_num in credit_rows:
    nettofalligkeit = sheet.cell(row_num, 16).value
    days_to_maturity = (nettofalligkeit - today_date).days
    credit_amount = sheet.cell(row_num, 18).value
    if days_to_maturity > 0:
        not_mature_credits += credit_amount
    elif 1 <= days_to_maturity <= 30:
        maturity_1_30_credits += credit_amount
    elif 31 <= days_to_maturity <= 60:
        maturity_31_60_credits += credit_amount
    else:
        mature_over_60_credits += credit_amount

# Calculate the percentage for each cluster
total_credit_amount = abs(float(credit_amount_sum))
percent_not_mature_credits = (abs(not_mature_credits) / total_credit_amount) * 100
percent_maturity_1_30_credits = (abs(maturity_1_30_credits) / total_credit_amount) * 100
percent_maturity_31_60_credits = (abs(maturity_31_60_credits) / total_credit_amount) * 100
percent_mature_over_60_credits = (abs(mature_over_60_credits) / total_credit_amount) * 100

# Output the ageing report for credit rows to the 'Analysis' sheet
analysis_sheet.append(['Ageing Report - Credits'])
analysis_sheet.append(['Cluster', 'Amount', 'Percentage'])
analysis_sheet.append(['Not Mature', not_mature_credits, percent_not_mature_credits])
analysis_sheet.append(['1-30 Days', maturity_1_30_credits, percent_maturity_1_30_credits])
analysis_sheet.append(['31-60 Days', maturity_31_60_credits, percent_maturity_31_60_credits])
analysis_sheet.append(['>60 Days', mature_over_60_credits, percent_mature_over_60_credits])

# ============
# Extract rows and amounts for credit positions
credit_positions = [(row, sheet.cell(row, 18).value) for row in credit_rows]

# Sort credit positions by amount (lowest to highest)
top_10_credits = sorted(credit_positions, key=lambda x: x[1])[:10]

# Output the top 10 credit positions to the 'Analysis' sheet
analysis_sheet.append(['Top 10 Credit Positions'])
for row_num, amount in top_10_credits:
    analysis_sheet.append([f'Row {row_num}', amount])

# ============
workbook.save(r"output/workbook_new.xlsx")